import csv
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import pathlib
import click
import tensorflow as tf
import numpy as np
import tqdm
import os

import config
from calc import truncate


def network(sess, hours):
    hours = list(map(lambda e: int(e), hours))
    hours = np.array(hours).reshape((1, len(hours)))

    graph = tf.get_default_graph()
    x = graph.get_tensor_by_name('Times:0')
    nn = graph.get_tensor_by_name('Output_layer/Tanh:0')

    marks = sess.run(nn, feed_dict={x: hours})[0]
    truncate(hours[0], marks)
    return marks


class Retriever:
    def __init__(self, network_session, output='output.xlsx'):
        self.__mark_cols = ['F', 'G', 'H', 'I', 'J', 'K', 'L']
        self.__mark_rows = [7, 52, 97, 142]
        self.__mark_total_col = 'O'
        self.__marks_sheet_name = 'Узагальнена Інформація '
        self.total_time = 0
        self.output = output
        self.sess = network_session

        self.color = '363636'
        self.row = 1
        self.book = self.__init_workbook()

    def process(self, file_path: pathlib.Path):
        file = file_path.as_posix()
        try:
            f = openpyxl.load_workbook(file, data_only=True)
            marks_ws = f[self.__marks_sheet_name]
            for v, mr in zip(self.__retrieve_hours(f), self.__mark_rows):
                total = float(marks_ws[f"{self.__mark_total_col}{mr}"].value)
                marks = network(self.sess, v)
                s = sum(marks)
                marks = [round(v / s * 100 * total) / 100 for v in marks]
                if abs(sum(marks)) > 0.1:
                    self.__append(file_path.name, mr, marks)
        except KeyError:
            pass
        except Exception as error:
            print(error)

    def save(self):
        self.book.save(self.output)

    def __init_workbook(self):
        wb = openpyxl.Workbook()
        ws: Worksheet = wb.active
        ws.title = 'Расчет оценок'
        ws.column_dimensions['A'].width = 34
        ws.column_dimensions['B'].width = 24

        self.__set_cell(ws, self.row, 1, 'Документ')
        self.__set_cell(ws, self.row, 2, 'Таблица')
        self.__set_cell(ws, self.row, 3, 'Строка')
        self.__set_cell(ws, self.row, 4, 'Посещения')
        self.__set_cell(ws, self.row, 5, 'Практика')
        self.__set_cell(ws, self.row, 6, 'Семинар')
        self.__set_cell(ws, self.row, 7, 'Лабор.')
        self.__set_cell(ws, self.row, 8, 'КРА')
        self.__set_cell(ws, self.row, 9, 'КРД')
        self.__set_cell(ws, self.row, 10, 'МКТ')
        self.__set_cell(ws, self.row, 11, 'Сумма')
        self.row += 1
        return wb

    def __append(self, name, row, marks):
        ws: Worksheet = self.book.active
        self.__set_cell(ws, self.row, 1, name)
        self.__set_cell(ws, self.row, 2, self.__marks_sheet_name)
        self.__set_cell(ws, self.row, 3, row)

        for c, v in zip(range(4, len(marks) + 4), marks):
            self.__set_cell(ws, self.row, c, v)
        self.__set_cell(ws, self.row, c + 1, sum(marks))
        self.row += 1

    def __set_cell(self, ws, row, column, value):
        cell = ws.cell(row, column)
        cell.value = value

    def __retrieve_marks(self, ws, row):
        ml = []
        t = float(ws[f"{self.__mark_total_col}{row}"].value)
        if t == 0:
            return None
        for c in self.__mark_cols:
            ml.append(round(float(ws[f"{c}{row}"].value)/t * 10_000) / 10_000)
        return ml

    def __retrieve_hours(self, wb):
        ws = wb['Робоча програма']
        pat = "Всього годин модуля"
        for r in ws:
            if r[0].value == pat:
                ind = r[0].row
                yield [ws[f"G{ind}"].value, ws[f"H{ind}"].value, ws[f"I{ind}"].value, ws[f"J{ind}"].value,
                       ws[f"K{ind}"].value, ws[f"L{ind}"].value]

    def __convert_time(self, sec):
        h, s = divmod(sec, 3600)
        m, s = divmod(s, 60)
        return f"{h}h {m}m {s:.2f}s"


def iter_through(dir_: pathlib.Path):
    """
    Function that goes recursively through ``dir_``. If ``dir_'s`` value is
    directory it goes through its files.

    :param dir_: ``pathlib.Path`` instance
    """
    for v in dir_.iterdir():
        if v.is_dir():
            yield from iter_through(v)
        else:
            yield v


@click.command()
@click.argument('in_')
@click.option('--out', default='processed', help='Specify output directory. Otherwise it will save into processed/ directory')
def run(in_, out):
    """
    Retrieve data from xlsx files that are laid in given directory.
    First argument is .xlsx files directory; second is the directory with file name for file to output data in.

    It uses pool with 4 processes to analyze files; so be careful when doing ``KeyboardInterrupt``

    *******
    Example
    *******

    ``python retriever.py c:/container/ d:/data.csv``

    """
    if not os.path.exists(os.path.dirname(out)):
        os.makedirs(os.path.dirname(out))
    try:
        with tf.Session() as sess:
            saver = tf.train.import_meta_graph(f'{config.MODEL_PATH}.meta')
            saver.restore(sess, config.MODEL_PATH)
            retriever = Retriever(output=out, network_session=sess)
            for v in tqdm.tqdm(list(iter_through(pathlib.Path(in_)))):
                retriever.process(v)
            retriever.save()
        print(f"Data processed")
    except KeyboardInterrupt:
        pass


if __name__ == '__main__':
    run()