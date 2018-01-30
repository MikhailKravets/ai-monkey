"""
Tool for retrieving data from special ``*.xls`` or ``*.xlsx`` file.
Also, it is possible to process directories of files.
It does generate special *.csv file with statistically important data.
"""
import csv

import openpyxl
import pathlib
from multiprocessing import Pool, Queue, Manager
import time

import click


def __retrieve_marks(ws, row):
    ml = []
    __mark_cols = ['F', 'G', 'H', 'I', 'J', 'K', 'L']
    __mark_total_col = 'O'
    t = float(ws[f"{__mark_total_col}{row}"].value)
    if t == 0:
        return None
    for c in __mark_cols:
        ml.append(round(float(ws[f"{c}{row}"].value) / t * 10_000) / 10_000)
    return ml


def __retrieve_hours(wb):
    ws = wb['Робоча програма']
    pat = "Всього годин модуля"
    for r in ws:
        if r[0].value == pat:
            ind = r[0].row
            yield [ws[f"G{ind}"].value, ws[f"H{ind}"].value, ws[f"I{ind}"].value, ws[f"J{ind}"].value,
                   ws[f"K{ind}"].value, ws[f"L{ind}"].value]

    
def process(file_name, queue: Queue):
    __mark_rows = [7, 52, 97, 142]
    print(f"Processing: {file_name}")
    try:
        f = openpyxl.load_workbook(file_name, data_only=True)
        marks_ws = f['Узагальнена Інформація ']
        for r in __mark_rows:
            marks_list = __retrieve_marks(marks_ws, r)
            if marks_list is not None:
                queue.put({
                    'marks': marks_list,
                    'hours': __retrieve_hours(f).__next__()
                })
    except KeyError:
        pass
    except Exception as error:
        pass
    print(f"Processed: {file_name}")


def save(output, queue: Queue):
    with open(output, "w") as csvfile:
        writer = csv.writer(csvfile)
        while not queue.empty():
            v = queue.get()
            writer.writerow(v['hours'] + ['-'] + v['marks'])


def iter_through(dir_: pathlib.Path):
    """
    Function that goes recursively through ``dir_``. If ``dir_'s`` value is
    directory it goes through its files.

    :param dir_: ``pathlib.Path`` instance
    """
    for v in dir_.iterdir():
        if v.is_dir():
            yield from iter_through(v)
        else:
            yield v


@click.command()
@click.argument('in_')
@click.option('--out', default='data.csv', help='Specify output directory. Otherwise it will save into data.csv')
def run(in_, out):
    """
    Retrieve data from xlsx files that are laid in given directory.
    First argument is .xlsx files directory; second is the directory for file to output data in.

    It uses pool with 4 processes to analyze files; so be careful when doing ``KeyboardInterrupt``

    *******
    Example
    *******

    ``python retriever.py c:/container/ d:/data.csv``

    """
    paths = []
    try:
        for v in iter_through(pathlib.Path(in_)):
            paths.append(str(v.absolute()))
        print(f"Begin to retrieve data from {in_}")
        with Pool(4) as p:
            q = Manager().Queue()
            p.starmap_async(process, [[v, q] for v in paths]).get(9999)
        save(out, q)
        print(f"Data saved to {out}")
    except KeyboardInterrupt:
        save(out, q)


if __name__ == '__main__':
    run()

